"""XLSX export builder with summary and widget sheets."""

import base64
from datetime import datetime
import re
from io import BytesIO
from typing import Any, Iterable, List, Set

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from export.metadata_rows import build_export_metadata_rows
from export.models import DashboardExportSnapshot, MAX_WIDGET_ROWS


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name or "Sheet").strip()
    return (cleaned or "Sheet")[:31]


def resolve_xlsx_timestamp(raw_timestamp: str) -> str:
    value = (raw_timestamp or "").strip()
    if value:
        try:
            normalized = value.replace("Z", "+00:00")
            return datetime.fromisoformat(normalized).strftime("%Y%m%d_%H%M%S")
        except ValueError:
            pass
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _sheet_prefix(widget_type: str, has_chart_data: bool) -> str:
    normalized = (widget_type or "").lower()
    if normalized == "kpi":
        return "KPI"
    if normalized == "table":
        return "Table"
    if has_chart_data:
        return "Chart"
    return "Data"


def _unique_sheet_name(base_name: str, used: Set[str]) -> str:
    safe = _safe_sheet_name(base_name)
    if safe not in used:
        used.add(safe)
        return safe

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = _safe_sheet_name(f"{safe[:31-len(suffix)]}{suffix}")
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def _append_rows(ws: Any, headers: Iterable[Any], rows: List[List[Any]], truncate_note: bool = True) -> None:
    ws.append(list(headers))
    capped = rows[:MAX_WIDGET_ROWS]
    for row in capped:
        ws.append(list(row))
    if truncate_note and len(rows) > len(capped):
        ws.append([f"TRUNCATED: Showing first {len(capped)} rows"])


def _decode_chart_image(raw: str) -> BytesIO | None:
    if not raw:
        return None
    value = raw.split(",", 1)[1] if "," in raw else raw
    try:
        binary = base64.b64decode(value)
        stream = BytesIO(binary)
        stream.seek(0)
        return stream
    except Exception:
        return None


def build_xlsx(snapshot: DashboardExportSnapshot) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    used_sheet_names: Set[str] = {"Summary"}

    for key, value in build_export_metadata_rows(snapshot, include_widget_count=True):
        summary.append([key, value])

    if not snapshot.widgets:
        info = wb.create_sheet(title="Info")
        info.append(["Info", "No widgets available for XLSX export"])

    for idx, widget in enumerate(snapshot.widgets, start=1):
        data = widget.data or {}
        has_chart_data = bool(data.get("labels") and data.get("datasets"))
        prefix = _sheet_prefix(widget.type, has_chart_data)
        base_title = f"{prefix}_{widget.title or 'UntitledWidget'}"
        ws = wb.create_sheet(title=_unique_sheet_name(base_title, used_sheet_names))

        if widget.type == "table":
            _append_rows(ws, data.get("headers", []), data.get("rows", []))
        elif data.get("labels") and data.get("datasets"):
            labels = data.get("labels", [])
            datasets = data.get("datasets", [])
            headers = ["Label", *[ds.get("label", "Value") for ds in datasets]]
            rows = []
            for i, label in enumerate(labels[:MAX_WIDGET_ROWS]):
                row = [label]
                for ds in datasets:
                    vals = ds.get("data", [])
                    row.append(vals[i] if i < len(vals) else "")
                rows.append(row)
            _append_rows(ws, headers, rows, truncate_note=False)
        elif widget.type == "kpi":
            _append_rows(
                ws,
                ["Metric", "Value"],
                [
                    [widget.title, data.get("value", "")],
                    ["Unit", data.get("unit", "")],
                    ["Change", data.get("change", "")],
                    ["Trend", data.get("trend", "")],
                    ["Subtitle", data.get("subtitle", "")],
                ],
                truncate_note=False,
            )
        else:
            ws.append(["Info", "No tabular representation available"])

        image_stream = _decode_chart_image(widget.chart_image_base64 or "")
        if image_stream is not None:
            try:
                image = XLImage(image_stream)
                image.width = 640
                image.height = 360
                ws.add_image(image, "H2")
            except Exception:
                ws.append([])
                ws.append(["Warning", "Chart image could not be embedded"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

